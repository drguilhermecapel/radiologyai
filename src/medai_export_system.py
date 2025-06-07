# export_system.py - Sistema de exportação e interoperabilidade

import numpy as np
import pandas as pd
from pathlib import Path
import json
import yaml
import h5py
import pickle
import zipfile
import tarfile
from typing import Dict, List, Optional, Any, Union
import pydicom
from pydicom.dataset import FileDataset
import nibabel as nib
import SimpleITK as sitk
import onnx
import onnxruntime as ort
import tensorflow as tf
from datetime import datetime
import logging
import base64
from cryptography.fernet import Fernet
import hashlib
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference, LineChart
import matplotlib.pyplot as plt
from io import BytesIO
import requests

logger = logging.getLogger('MedAI.Export')

class ExportManager:
    """
    Gerenciador de exportação e interoperabilidade
    Suporta múltiplos formatos e padrões médicos
    """
    
    def __init__(self, encryption_key: Optional[str] = None):
        self.supported_formats = {
            'model': ['h5', 'pb', 'onnx', 'tflite', 'pt', 'pkl'],
            'image': ['dcm', 'nii', 'nii.gz', 'png', 'jpg', 'tiff', 'nrrd'],
            'data': ['csv', 'xlsx', 'json', 'parquet', 'feather', 'hdf5'],
            'report': ['pdf', 'html', 'docx', 'tex', 'md', 'xml'],
            'archive': ['zip', 'tar', 'tar.gz', 'tar.bz2']
        }
        
        # Configurar criptografia se fornecida
        if encryption_key:
            self.fernet = Fernet(encryption_key.encode() if isinstance(encryption_key, str) else encryption_key)
        else:
            self.fernet = None
            
        logger.info("ExportManager inicializado")
    
    def export_model(self,
                    model: Any,
                    output_path: str,
                    format: str = 'h5',
                    include_metadata: bool = True,
                    optimize: bool = False) -> Dict[str, Any]:
        """
        Exporta modelo para diferentes formatos
        
        Args:
            model: Modelo a exportar
            output_path: Caminho de saída
            format: Formato de exportação
            include_metadata: Se deve incluir metadados
            optimize: Se deve otimizar o modelo
            
        Returns:
            Informações da exportação
        """
        output_path = Path(output_path)
        export_info = {
            'format': format,
            'path': str(output_path),
            'timestamp': datetime.now().isoformat(),
            'size_bytes': 0,
            'metadata_included': include_metadata,
            'optimized': optimize
        }
        
        try:
            if format == 'h5':
                # Keras/TensorFlow H5
                model.save(output_path)
                
            elif format == 'pb':
                # TensorFlow SavedModel
                tf.saved_model.save(model, str(output_path.parent / output_path.stem))
                
            elif format == 'onnx':
                # ONNX
                self._export_to_onnx(model, output_path, optimize)
                
            elif format == 'tflite':
                # TensorFlow Lite
                self._export_to_tflite(model, output_path, optimize)
                
            elif format == 'pt':
                # PyTorch
                if hasattr(model, 'state_dict'):
                    import torch
                    torch.save(model.state_dict(), output_path)
                else:
                    raise ValueError("Modelo não é PyTorch")
                    
            elif format == 'pkl':
                # Pickle genérico
                with open(output_path, 'wb') as f:
                    pickle.dump(model, f)
                    
            else:
                raise ValueError(f"Formato não suportado: {format}")
            
            # Adicionar metadados se solicitado
            if include_metadata:
                metadata = self._generate_model_metadata(model)
                metadata_path = output_path.with_suffix('.meta.json')
                with open(metadata_path, 'w') as f:
                    json.dump(metadata, f, indent=2)
                export_info['metadata_path'] = str(metadata_path)
            
            # Calcular tamanho
            if output_path.exists():
                export_info['size_bytes'] = output_path.stat().st_size
            elif output_path.is_dir():
                export_info['size_bytes'] = sum(
                    f.stat().st_size for f in output_path.rglob('*') if f.is_file()
                )
            
            # Calcular hash para verificação
            export_info['sha256'] = self._calculate_file_hash(output_path)
            
            logger.info(f"Modelo exportado: {output_path} ({format})")
            
        except Exception as e:
            logger.error(f"Erro ao exportar modelo: {str(e)}")
            export_info['error'] = str(e)
            raise
            
        return export_info
    
    def _export_to_onnx(self, model: tf.keras.Model, output_path: Path, optimize: bool):
        """Exporta modelo TensorFlow para ONNX"""
        import tf2onnx
        
        # Especificar assinatura de entrada
        spec = (tf.TensorSpec(model.input_shape, tf.float32, name="input"),)
        
        # Converter
        model_proto, _ = tf2onnx.convert.from_keras(
            model, 
            input_signature=spec,
            opset=13,
            output_path=str(output_path)
        )
        
        if optimize:
            # Otimizar modelo ONNX
            from onnxruntime.transformers import optimizer
            optimized_model = optimizer.optimize_model(str(output_path))
            optimized_model.save_model_to_file(str(output_path))
    
    def _export_to_tflite(self, model: tf.keras.Model, output_path: Path, optimize: bool):
        """Exporta modelo para TensorFlow Lite"""
        converter = tf.lite.TFLiteConverter.from_keras_model(model)
        
        if optimize:
            # Otimizações
            converter.optimizations = [tf.lite.Optimize.DEFAULT]
            # Quantização dinâmica
            converter.target_spec.supported_types = [tf.float16]
        
        tflite_model = converter.convert()
        
        with open(output_path, 'wb') as f:
            f.write(tflite_model)
    
    def _generate_model_metadata(self, model: Any) -> Dict[str, Any]:
        """Gera metadados do modelo"""
        metadata = {
            'export_timestamp': datetime.now().isoformat(),
            'framework': self._detect_framework(model),
            'architecture': {},
            'training_info': {},
            'performance_metrics': {}
        }
        
        # TensorFlow/Keras
        if hasattr(model, 'summary'):
            # Capturar summary
            import io
            stream = io.StringIO()
            model.summary(print_fn=lambda x: stream.write(x + '\n'))
            metadata['architecture']['summary'] = stream.getvalue()
            metadata['architecture']['num_parameters'] = model.count_params()
            metadata['architecture']['input_shape'] = str(model.input_shape)
            metadata['architecture']['output_shape'] = str(model.output_shape)
            
            # Configuração do modelo
            if hasattr(model, 'get_config'):
                metadata['architecture']['config'] = model.get_config()
        
        # PyTorch
        elif hasattr(model, 'state_dict'):
            metadata['architecture']['num_parameters'] = sum(
                p.numel() for p in model.parameters()
            )
            metadata['architecture']['layers'] = [
                str(module) for module in model.modules()
            ]
        
        return metadata
    
    def _detect_framework(self, model: Any) -> str:
        """Detecta framework do modelo"""
        if hasattr(model, 'keras'):
            return 'tensorflow'
        elif hasattr(model, 'state_dict') and hasattr(model, 'forward'):
            return 'pytorch'
        elif hasattr(model, 'predict') and hasattr(model, 'fit'):
            return 'sklearn'
        else:
            return 'unknown'
    
    def export_medical_image(self,
                           image: np.ndarray,
                           output_path: str,
                           format: str = 'dcm',
                           metadata: Optional[Dict] = None,
                           patient_info: Optional[Dict] = None) -> Dict[str, Any]:
        """
        Exporta imagem médica em vários formatos
        
        Args:
            image: Array da imagem
            output_path: Caminho de saída
            format: Formato de exportação
            metadata: Metadados da imagem
            patient_info: Informações do paciente
            
        Returns:
            Informações da exportação
        """
        output_path = Path(output_path)
        export_info = {
            'format': format,
            'path': str(output_path),
            'timestamp': datetime.now().isoformat()
        }
        
        try:
            if format == 'dcm':
                # DICOM
                self._export_to_dicom(image, output_path, metadata, patient_info)
                
            elif format in ['nii', 'nii.gz']:
                # NIfTI
                self._export_to_nifti(image, output_path, metadata)
                
            elif format == 'nrrd':
                # NRRD
                import nrrd
                header = metadata or {}
                nrrd.write(str(output_path), image, header)
                
            elif format in ['png', 'jpg', 'jpeg', 'tiff']:
                # Formatos de imagem padrão
                import cv2
                # Normalizar para 8-bit se necessário
                if image.dtype != np.uint8:
                    image_norm = ((image - image.min()) / (image.max() - image.min()) * 255).astype(np.uint8)
                else:
                    image_norm = image
                cv2.imwrite(str(output_path), image_norm)
                
            else:
                raise ValueError(f"Formato de imagem não suportado: {format}")
            
            export_info['size_bytes'] = output_path.stat().st_size
            export_info['image_shape'] = image.shape
            export_info['image_dtype'] = str(image.dtype)
            
            logger.info(f"Imagem exportada: {output_path} ({format})")
            
        except Exception as e:
            logger.error(f"Erro ao exportar imagem: {str(e)}")
            export_info['error'] = str(e)
            raise
            
        return export_info
    
    def _export_to_dicom(self, 
                        image: np.ndarray, 
                        output_path: Path,
                        metadata: Optional[Dict],
                        patient_info: Optional[Dict]):
        """Exporta imagem para formato DICOM"""
        from pydicom.dataset import Dataset, FileDataset
        from pydicom.uid import generate_uid
        
        # Criar FileDataset
        file_meta = Dataset()
        file_meta.MediaStorageSOPClassUID = pydicom.uid.SecondaryCaptureImageStorage
        file_meta.MediaStorageSOPInstanceUID = generate_uid()
        file_meta.TransferSyntaxUID = pydicom.uid.ExplicitVRLittleEndian
        
        ds = FileDataset(
            str(output_path), 
            {}, 
            file_meta=file_meta, 
            preamble=b"\0" * 128
        )
        
        # Informações obrigatórias
        ds.PatientName = patient_info.get('PatientName', 'Anonymous') if patient_info else 'Anonymous'
        ds.PatientID = patient_info.get('PatientID', '000000') if patient_info else '000000'
        
        # Data e hora
        dt = datetime.now()
        ds.StudyDate = dt.strftime('%Y%m%d')
        ds.StudyTime = dt.strftime('%H%M%S.%f')
        ds.ContentDate = ds.StudyDate
        ds.ContentTime = ds.StudyTime
        
        # UIDs
        ds.StudyInstanceUID = generate_uid()
        ds.SeriesInstanceUID = generate_uid()
        ds.SOPInstanceUID = file_meta.MediaStorageSOPInstanceUID
        ds.SOPClassUID = file_meta.MediaStorageSOPClassUID
        
        # Informações da imagem
        ds.Modality = metadata.get('Modality', 'OT') if metadata else 'OT'
        ds.ConversionType = 'WSD'
        ds.PhotometricInterpretation = 'MONOCHROME2'
        ds.SamplesPerPixel = 1
        ds.Rows, ds.Columns = image.shape[:2]
        
        # Pixel data
        if image.dtype != np.uint16:
            image = ((image - image.min()) / (image.max() - image.min()) * 65535).astype(np.uint16)
        
        ds.PixelData = image.tobytes()
        ds.BitsAllocated = 16
        ds.BitsStored = 16
        ds.HighBit = 15
        ds.PixelRepresentation = 0
        
        # Adicionar metadados extras se fornecidos
        if metadata:
            for key, value in metadata.items():
                if hasattr(ds, key):
                    setattr(ds, key, value)
        
        # Salvar
        ds.save_as(str(output_path), write_like_original=False)
    
    def _export_to_nifti(self, 
                        image: np.ndarray, 
                        output_path: Path,
                        metadata: Optional[Dict]):
        """Exporta imagem para formato NIfTI"""
        # Criar imagem NIfTI
        if len(image.shape) == 2:
            # Adicionar dimensão Z se 2D
            image = image[:, :, np.newaxis]
        
        # Affine padrão (identidade)
        affine = np.eye(4)
        if metadata and 'affine' in metadata:
            affine = metadata['affine']
        elif metadata and 'spacing' in metadata:
            spacing = metadata['spacing']
            affine[0, 0] = spacing[0]
            affine[1, 1] = spacing[1]
            affine[2, 2] = spacing[2] if len(spacing) > 2 else 1
        
        # Criar e salvar
        nifti_img = nib.Nifti1Image(image, affine)
        
        # Adicionar header info se disponível
        if metadata:
            header = nifti_img.header
            if 'description' in metadata:
                header['descrip'] = metadata['description'][:80]  # Limite de 80 chars
        
        nib.save(nifti_img, str(output_path))
    
    def export_data(self,
                   data: Union[pd.DataFrame, Dict, List],
                   output_path: str,
                   format: str = 'csv',
                   compression: Optional[str] = None,
                   encrypt: bool = False) -> Dict[str, Any]:
        """
        Exporta dados em vários formatos
        
        Args:
            data: Dados a exportar
            output_path: Caminho de saída
            format: Formato de exportação
            compression: Tipo de compressão
            encrypt: Se deve criptografar
            
        Returns:
            Informações da exportação
        """
        output_path = Path(output_path)
        export_info = {
            'format': format,
            'path': str(output_path),
            'timestamp': datetime.now().isoformat(),
            'compressed': compression is not None,
            'encrypted': encrypt
        }
        
        try:
            # Converter para DataFrame se necessário
            if not isinstance(data, pd.DataFrame):
                if isinstance(data, dict):
                    df = pd.DataFrame(data)
                elif isinstance(data, list):
                    df = pd.DataFrame(data)
                else:
                    raise ValueError("Tipo de dados não suportado")
            else:
                df = data
            
            # Exportar baseado no formato
            if format == 'csv':
                df.to_csv(output_path, index=False, compression=compression)
                
            elif format == 'xlsx':
                # Excel com formatação
                self._export_to_excel(df, output_path)
                
            elif format == 'json':
                df.to_json(output_path, orient='records', indent=2)
                
            elif format == 'parquet':
                df.to_parquet(output_path, compression=compression or 'snappy')
                
            elif format == 'feather':
                df.to_feather(output_path)
                
            elif format == 'hdf5':
                df.to_hdf(output_path, key='data', mode='w', complevel=9 if compression else 0)
                
            else:
                raise ValueError(f"Formato de dados não suportado: {format}")
            
            # Criptografar se solicitado
            if encrypt and self.fernet:
                self._encrypt_file(output_path)
                export_info['encrypted'] = True
            
            export_info['size_bytes'] = output_path.stat().st_size
            export_info['num_rows'] = len(df)
            export_info['num_columns'] = len(df.columns)
            
            logger.info(f"Dados exportados: {output_path} ({format})")
            
        except Exception as e:
            logger.error(f"Erro ao exportar dados: {str(e)}")
            export_info['error'] = str(e)
            raise
            
        return export_info
    
    def _export_to_excel(self, df: pd.DataFrame, output_path: Path):
        """Exporta DataFrame para Excel com formatação"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Escrever dados
            df.to_excel(writer, sheet_name='Dados', index=False)
            
            # Obter worksheet
            workbook = writer.book
            worksheet = writer.sheets['Dados']
            
            # Formatar cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Adicionar filtros
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Criar sheet de estatísticas
            stats_sheet = workbook.create_sheet('Estatísticas')
            
            # Estatísticas básicas
            stats_data = [
                ['Métrica', 'Valor'],
                ['Total de Registros', len(df)],
                ['Total de Colunas', len(df.columns)],
                ['Tamanho em Memória (bytes)', df.memory_usage(deep=True).sum()]
            ]
            
            # Adicionar estatísticas numéricas
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            for col in numeric_cols:
                stats_data.extend([
                    [f'{col} - Média', df[col].mean()],
                    [f'{col} - Desvio Padrão', df[col].std()],
                    [f'{col} - Min', df[col].min()],
                    [f'{col} - Max', df[col].max()]
                ])
            
            for row_idx, row_data in enumerate(stats_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    stats_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Formatar sheet de estatísticas
            for cell in stats_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
    
    def export_report(self,
                     content: Dict[str, Any],
                     output_path: str,
                     format: str = 'pdf',
                     template: Optional[str] = None,
                     style: str = 'clinical') -> Dict[str, Any]:
        """
        Exporta relatório em vários formatos
        
        Args:
            content: Conteúdo do relatório
            output_path: Caminho de saída
            format: Formato de exportação
            template: Template a usar
            style: Estilo do relatório
            
        Returns:
            Informações da exportação
        """
        output_path = Path(output_path)
        export_info = {
            'format': format,
            'path': str(output_path),
            'timestamp': datetime.now().isoformat(),
            'style': style
        }
        
        try:
            if format == 'pdf':
                self._export_report_pdf(content, output_path, style)
                
            elif format == 'html':
                self._export_report_html(content, output_path, template)
                
            elif format == 'docx':
                self._export_report_docx(content, output_path)
                
            elif format == 'tex':
                self._export_report_latex(content, output_path)
                
            elif format == 'md':
                self._export_report_markdown(content, output_path)
                
            elif format == 'xml':
                self._export_report_xml(content, output_path)
                
            else:
                raise ValueError(f"Formato de relatório não suportado: {format}")
            
            export_info['size_bytes'] = output_path.stat().st_size
            
            logger.info(f"Relatório exportado: {output_path} ({format})")
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório: {str(e)}")
            export_info['error'] = str(e)
            raise
            
        return export_info
    
    def _export_report_pdf(self, content: Dict, output_path: Path, style: str):
        """Exporta relatório para PDF"""
        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = styles['Title']
        story.append(Paragraph(content.get('title', 'Relatório Médico'), title_style))
        story.append(Spacer(1, 12))
        
        # Metadados
        if 'metadata' in content:
            metadata_data = [
                ['Campo', 'Valor']
            ]
            for key, value in content['metadata'].items():
                metadata_data.append([key, str(value)])
            
            metadata_table = Table(metadata_data)
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metadata_table)
            story.append(Spacer(1, 12))
        
        # Seções
        for section in content.get('sections', []):
            # Título da seção
            story.append(Paragraph(section.get('title', ''), styles['Heading1']))
            
            # Conteúdo da seção
            if 'text' in section:
                story.append(Paragraph(section['text'], styles['BodyText']))
                story.append(Spacer(1, 12))
            
            # Tabela se houver
            if 'table' in section:
                table = Table(section['table'])
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ]))
                story.append(table)
                story.append(Spacer(1, 12))
            
            # Imagem se houver
            if 'image' in section:
                # Decodificar imagem base64 se necessário
                image_data = section['image']
                if isinstance(image_data, str) and image_data.startswith('data:image'):
                    # Extrair dados base64
                    header, data = image_data.split(',', 1)
                    image_bytes = base64.b64decode(data)
                    
                    # Criar imagem temporária
                    from reportlab.platypus import Image
                    from io import BytesIO
                    img = Image(BytesIO(image_bytes), width=400, height=300)
                    story.append(img)
                    story.append(Spacer(1, 12))
        
        # Construir PDF
        doc.build(story)
    
    def _export_report_html(self, content: Dict, output_path: Path, template: Optional[str]):
        """Exporta relatório para HTML"""
        # Template HTML
        if template:
            with open(template, 'r') as f:
                html_template = f.read()
        else:
            html_template = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        .metadata {{ background-color: #f0f0f0; padding: 10px; margin: 10px 0; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    {content}
</body>
</html>
"""
        
        # Construir conteúdo HTML
        html_content = ""
        
        # Metadados
        if 'metadata' in content:
            html_content += '<div class="metadata">'
            for key, value in content['metadata'].items():
                html_content += f'<p><strong>{key}:</strong> {value}</p>'
            html_content += '</div>'
        
        # Seções
        for section in content.get('sections', []):
            html_content += f'<h2>{section.get("title", "")}</h2>'
            
            if 'text' in section:
                html_content += f'<p>{section["text"]}</p>'
            
            if 'table' in section:
                html_content += '<table>'
                for i, row in enumerate(section['table']):
                    html_content += '<tr>'
                    tag = 'th' if i == 0 else 'td'
                    for cell in row:
                        html_content += f'<{tag}>{cell}</{tag}>'
                    html_content += '</tr>'
                html_content += '</table>'
        
        # Substituir no template
        html = html_template.format(
            title=content.get('title', 'Relatório'),
            content=html_content
        )
        
        # Salvar
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
    
    def _export_report_markdown(self, content: Dict, output_path: Path):
        """Exporta relatório para Markdown"""
        md_content = f"# {content.get('title', 'Relatório')}\n\n"
        
        # Metadados
        if 'metadata' in content:
            md_content += "## Metadados\n\n"
            for key, value in content['metadata'].items():
                md_content += f"- **{key}**: {value}\n"
            md_content += "\n"
        
        # Seções
        for section in content.get('sections', []):
            md_content += f"## {section.get('title', '')}\n\n"
            
            if 'text' in section:
                md_content += f"{section['text']}\n\n"
            
            if 'table' in section:
                # Cabeçalho
                if section['table']:
                    md_content += "| " + " | ".join(str(cell) for cell in section['table'][0]) + " |\n"
                    md_content += "|" + "---|" * len(section['table'][0]) + "\n"
                    
                    # Linhas
                    for row in section['table'][1:]:
                        md_content += "| " + " | ".join(str(cell) for cell in row) + " |\n"
                    md_content += "\n"
        
        # Salvar
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md_content)
    
    def create_archive(self,
                      files: List[Union[str, Path]],
                      output_path: str,
                      format: str = 'zip',
                      compression: str = 'default',
                      encrypt: bool = False,
                      password: Optional[str] = None) -> Dict[str, Any]:
        """
        Cria arquivo compactado com múltiplos arquivos
        
        Args:
            files: Lista de arquivos para incluir
            output_path: Caminho do arquivo de saída
            format: Formato do arquivo (zip, tar, etc)
            compression: Tipo de compressão
            encrypt: Se deve criptografar
            password: Senha para arquivos zip
            
        Returns:
            Informações do arquivo criado
        """
        output_path = Path(output_path)
        archive_info = {
            'format': format,
            'path': str(output_path),
            'timestamp': datetime.now().isoformat(),
            'num_files': len(files),
            'compressed': True,
            'encrypted': encrypt
        }
        
        try:
            if format == 'zip':
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    if password:
                        zf.setpassword(password.encode())
                    
                    for file in files:
                        file_path = Path(file)
                        if file_path.exists():
                            arcname = file_path.name
                            zf.write(file_path, arcname)
                            logger.info(f"Adicionado ao arquivo: {arcname}")
                        else:
                            logger.warning(f"Arquivo não encontrado: {file_path}")
            
            elif format in ['tar', 'tar.gz', 'tar.bz2']:
                mode = 'w'
                if format == 'tar.gz':
                    mode = 'w:gz'
                elif format == 'tar.bz2':
                    mode = 'w:bz2'
                
                with tarfile.open(output_path, mode) as tf:
                    for file in files:
                        file_path = Path(file)
                        if file_path.exists():
                            arcname = file_path.name
                            tf.add(file_path, arcname)
                            logger.info(f"Adicionado ao arquivo: {arcname}")
                        else:
                            logger.warning(f"Arquivo não encontrado: {file_path}")
            
            else:
                raise ValueError(f"Formato de arquivo não suportado: {format}")
            
            # Criptografar arquivo completo se solicitado
            if encrypt and self.fernet and not password:
                self._encrypt_file(output_path)
            
            archive_info['size_bytes'] = output_path.stat().st_size
            archive_info['compression_ratio'] = self._calculate_compression_ratio(files, output_path)
            
            logger.info(f"Arquivo criado: {output_path}")
            
        except Exception as e:
            logger.error(f"Erro ao criar arquivo: {str(e)}")
            archive_info['error'] = str(e)
            raise
            
        return archive_info
    
    def _encrypt_file(self, file_path: Path):
        """Criptografa arquivo usando Fernet"""
        if not self.fernet:
            raise ValueError("Chave de criptografia não configurada")
        
        with open(file_path, 'rb') as f:
            data = f.read()
        
        encrypted_data = self.fernet.encrypt(data)
        
        encrypted_path = file_path.with_suffix(file_path.suffix + '.enc')
        with open(encrypted_path, 'wb') as f:
            f.write(encrypted_data)
        
        # Remover arquivo original
        file_path.unlink()
        
        # Renomear arquivo criptografado
        encrypted_path.rename(file_path)
    
    def _calculate_file_hash(self, file_path: Path) -> str:
        """Calcula hash SHA256 do arquivo"""
        sha256_hash = hashlib.sha256()
        
        with open(file_path, 'rb') as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        
        return sha256_hash.hexdigest()
    
    def _calculate_compression_ratio(self, original_files: List[Path], compressed_file: Path) -> float:
        """Calcula taxa de compressão"""
        original_size = sum(Path(f).stat().st_size for f in original_files if Path(f).exists())
        compressed_size = compressed_file.stat().st_size
        
        if original_size > 0:
            return 1 - (compressed_size / original_size)
        return 0.0
    
    def export_to_cloud(self,
                       file_path: str,
                       service: str,
                       credentials: Dict[str, str],
                       destination: str) -> Dict[str, Any]:
        """
        Exporta arquivo para serviço de nuvem
        
        Args:
            file_path: Arquivo local para upload
            service: Serviço de nuvem (aws, gcp, azure)
            credentials: Credenciais de acesso
            destination: Caminho de destino na nuvem
            
        Returns:
            Informações do upload
        """
        upload_info = {
            'service': service,
            'source': file_path,
            'destination': destination,
            'timestamp': datetime.now().isoformat()
        }
        
        try:
            if service == 'aws':
                import boto3
                
                s3 = boto3.client(
                    's3',
                    aws_access_key_id=credentials['access_key'],
                    aws_secret_access_key=credentials['secret_key']
                )
                
                bucket, key = destination.split('/', 1)
                s3.upload_file(file_path, bucket, key)
                
                upload_info['url'] = f"s3://{bucket}/{key}"
                
            elif service == 'gcp':
                from google.cloud import storage
                
                client = storage.Client.from_service_account_json(
                    credentials['service_account_json']
                )
                
                bucket_name, blob_name = destination.split('/', 1)
                bucket = client.bucket(bucket_name)
                blob = bucket.blob(blob_name)
                
                blob.upload_from_filename(file_path)
                
                upload_info['url'] = f"gs://{bucket_name}/{blob_name}"
                
            elif service == 'azure':
                from azure.storage.blob import BlobServiceClient
                
                blob_service = BlobServiceClient(
                    account_url=credentials['account_url'],
                    credential=credentials['access_key']
                )
                
                container, blob_name = destination.split('/', 1)
                blob_client = blob_service.get_blob_client(
                    container=container,
                    blob=blob_name
                )
                
                with open(file_path, 'rb') as data:
                    blob_client.upload_blob(data, overwrite=True)
                
                upload_info['url'] = blob_client.url
                
            else:
                raise ValueError(f"Serviço de nuvem não suportado: {service}")
            
            upload_info['success'] = True
            logger.info(f"Upload concluído: {destination}")
            
        except Exception as e:
            logger.error(f"Erro no upload: {str(e)}")
            upload_info['error'] = str(e)
            upload_info['success'] = False
            raise
            
        return upload_info
    
    def validate_export(self, export_info: Dict[str, Any]) -> bool:
        """
        Valida integridade da exportação
        
        Args:
            export_info: Informações da exportação
            
        Returns:
            True se válido
        """
        try:
            file_path = Path(export_info['path'])
            
            # Verificar se arquivo existe
            if not file_path.exists():
                logger.error(f"Arquivo não encontrado: {file_path}")
                return False
            
            # Verificar tamanho
            actual_size = file_path.stat().st_size
            if 'size_bytes' in export_info:
                if actual_size != export_info['size_bytes']:
                    logger.error(f"Tamanho incorreto: esperado {export_info['size_bytes']}, atual {actual_size}")
                    return False
            
            # Verificar hash se disponível
            if 'sha256' in export_info:
                actual_hash = self._calculate_file_hash(file_path)
                if actual_hash != export_info['sha256']:
                    logger.error("Hash não corresponde")
                    return False
            
            logger.info(f"Exportação validada: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Erro na validação: {str(e)}")
            return False
